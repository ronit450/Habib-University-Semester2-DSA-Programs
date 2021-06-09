import openpyxl
import os

# absolute directory path to find the directory of sorted xl file.


def AbsolutePath(dir):
    AbsoluteDirectory = os.path.dirname(os.path.abspath(__file__))
    returnDir = os.path.join(AbsoluteDirectory, dir)
    return returnDir

# Performs linear search in front and backwards direction based on the value of the default parameter
# if backwards is set as true, the linear search in backwards direction is performed
# index takes value from BinarySearch, element ios the area we are looking for in XL file


def LinearSearch(XL, index, element, backwards=False):
    # nextElem checks if some different value has been reached, indicating that the search needs to end
    nextElem = False
    if backwards:
        while not(nextElem) and index != 2:
            index -= 1
            StrCounter = 'E'+str(index)
            if XL[StrCounter].value != element:
                nextElem = True
                index += 1
    else:
        while not(nextElem) and index != XL.max_row:
            index += 1
            StrCounter = 'E'+str(index)
            if XL[StrCounter].value != element:
                nextElem = True
                index -= 1
    return index

# Binary function which will give the index of student record matching with the area


def BinarySearch(xl, element, lowerbound, upperbound):
    if lowerbound > upperbound:
        # returns -1 if no student is found
        return -1
    MidIndex = (lowerbound + upperbound) // 2
    StrCounter = 'E'+str(MidIndex)
    if element > xl[StrCounter].value:
        lowerbound = MidIndex + 1
        return BinarySearch(xl, element, lowerbound, upperbound)
    elif element < xl[StrCounter].value:
        upperbound = MidIndex - 1
        return BinarySearch(xl, element, lowerbound, upperbound)
    return MidIndex


# returns the recors of every student matching the given Area
# lowerindex and upperindex values are taken from the LinearSearch() function
def getStudentsRecord(xl, lowerindex, upperindex):
    Record = []
    SN = 0
    for i in range(lowerindex, upperindex+1):
        StrCounter = str(i)
        FName = xl['A' + StrCounter].value
        LName = xl['B' + StrCounter].value
        StdID = xl['C' + StrCounter].value
        Email = xl['D' + StrCounter].value
        recordTuple = (SN, FName, LName, StdID, Email)
        Record.append(recordTuple)
    return Record


# Area stored in parameter element is looked upon in the XL file
# all the previously defined helper functions are then used to get the student records from the xl file

def BinaryLinearSearch(xl, element):
    Lowerbound = 2
    Upperbound = xl.max_row
    # Sending the data to function BinarySearch which will search the area with binary search and return the index
    # where the area is found
    Mid = BinarySearch(xl, element, Lowerbound, Upperbound)
    if Mid != -1:
        # after that index it will be sent to linear search function where the index will be traversed linearly
        # upwards and downwards until anyother area is found.
        LowerIndex = LinearSearch(xl, Mid, element, True)
        UpperIndex = LinearSearch(xl, Mid, element)
        return getStudentsRecord(xl, LowerIndex, UpperIndex)
    # returns -1 if not any single student live in your area
    return -1


# in this main function, we will input area. The function will run Binary Linear Search on the excel file and get
# the students record from that area will be returned.

def main(Area):
    Database = openpyxl.load_workbook(
        filename=AbsolutePath('SortedDatabase.xlsx'))
    Datasheet = Database.active
    Records = BinaryLinearSearch(Datasheet, Area)
    if Records == -1:
        return 'No student in your Area'
    return Records
