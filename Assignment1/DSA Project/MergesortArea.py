import openpyxl
from openpyxl import Workbook

# This fucntion will read the excel file and then seperate student data. Store each student's data in tuple
# and store it in a list. The function will return the list.


def readXLSX():
    recordList = []
    Database = openpyxl.load_workbook(filename='Database.xlsx')
    Datasheet = Database.active
    for i in range(2, Datasheet.max_row+1):
        StrCounter = str(i)
        FName = Datasheet['A' + StrCounter].value
        LName = Datasheet['B' + StrCounter].value
        StdID = Datasheet['C' + StrCounter].value
        Email = Datasheet['D' + StrCounter].value
        Area = Datasheet['E' + StrCounter].value
        recordTuple = (FName, LName, StdID, Email, Area)
        recordList.append(recordTuple)
    return recordList


# This is a merge sort's merge function where the data will be merged according to its column.

def merge(left, right, lst, col):
    LeftIterator = 0
    RightIterator = 0
    MainIterator = 0
    while LeftIterator < len(left) and RightIterator < len(right):
        Condition = left[LeftIterator][col] <= right[RightIterator][col]
        if Condition:
            lst[MainIterator] = left[LeftIterator]
            LeftIterator += 1
        else:
            lst[MainIterator] = right[RightIterator]
            RightIterator += 1
        MainIterator += 1
    if LeftIterator < len(left) and RightIterator >= len(right):
        for i in range(LeftIterator, len(left)):
            lst[MainIterator] = left[i]
            MainIterator += 1
    elif RightIterator < len(right) and LeftIterator >= len(left):
        for i in range(RightIterator, len(right)):
            lst[MainIterator] = right[i]
            MainIterator += 1
    return lst

# Main Merge Sort function which takes parameter lst which is the list of tuples and col is the column of tuple upon which the lst needs
# to be sorted


def mergeSort(lst, col):
    if len(lst) != 1:
        MidIndex = len(lst) // 2
        Left = lst[:MidIndex]
        Right = lst[MidIndex:]
        mergeSort(Left, col)
        mergeSort(Right, col)
        lst = merge(Left, Right, lst, col)
    # returns sorted list
    return lst


# this function will create a new excel file and then assign the data of students sorted by the area.

def main():
    Newworkbook = Workbook()
    Newsheet = Newworkbook.active
    # sorting the excel file on the basis of column 4 which is area column
    # the first parameter is the function readXLSX which will read the excel file and returns the records of student as list
    SortedList = mergeSort(readXLSX(), 4)
    # assigning columns their name
    Newsheet['A1'] = 'First Name'
    Newsheet['B1'] = 'Last Name'
    Newsheet['C1'] = 'Habib Id'
    Newsheet['D1'] = 'Email Id'
    Newsheet['E1'] = 'Area'
    Counter = 2
    for i in SortedList:
        StrCounter = str(Counter)
        Newsheet['A' + StrCounter] = i[0]
        Newsheet['B' + StrCounter] = i[1]
        Newsheet['C' + StrCounter] = i[2]
        Newsheet['D' + StrCounter] = i[3]
        Newsheet['E' + StrCounter] = i[4]
        Counter += 1
    Newworkbook.save(filename='SortedDatabase.xlsx')


# main()  # Un_comment to re-create SortedDatabase file
