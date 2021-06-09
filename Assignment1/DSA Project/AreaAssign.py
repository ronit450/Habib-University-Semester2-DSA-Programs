import openpyxl
import random

# loads the already existing DSASurvey.xlsx
# read the xl file and create the list of areas present with repitation
# repitation is kept to retain the probability of an area being randomly assigned for the sake of beta project


def readArea():
    Area = openpyxl.load_workbook(filename='DSASurvey.xlsx')
    AreaSheet = Area.active
    AreaList = []
    for i in range(2, AreaSheet.max_row+1):
        temp = 'B' + str(i)
        AreaList.append(AreaSheet[temp].value)
    return AreaList

# Passed as the parameter is the area list we get from readArea() function
# uses this list to randomly return an element from the list


def getRandomArea(AL):
    Upperbound = len(AL)-1
    Lowerbound = 0
    RandomIndex = random.randint(Lowerbound, Upperbound)
    return AL[RandomIndex]

# Opens the pre-existing Data.xlsx file
# creates an additional column named Area
# randomly assign area to every student and save the file as Database.xlsx


def main():
    Database = openpyxl.load_workbook(filename='Data.xlsx')
    MainSheet = Database.active
    MainSheet['E1'] = 'Area'
    for i in range(2, MainSheet.max_row+1):
        Col = 'E' + str(i)
        MainSheet[Col] = getRandomArea(readArea())
    Database.save(filename='Database.xlsx')


# main() #Un-comment to run it again and re-create Database.xlsx file
