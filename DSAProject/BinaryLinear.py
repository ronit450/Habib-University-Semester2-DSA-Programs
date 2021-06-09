import openpyxl
import os


def AbsolutePath(dir):
    AbsoluteDirectory = os.path.dirname(os.path.abspath(__file__))
    returnDir = os.path.join(AbsoluteDirectory, dir)
    return returnDir


def LinearSearch(XL, index, element, backwards=False):
    nextElem = False
    if backwards:
        while not(nextElem) and index != 2:
            index -= 1
            StrCounter = 'E'+str(index)
            if XL[StrCounter].value != element:
                nextElem = True
                index += 1
    else:
        while not(nextElem) and index != XL.max_row:
            index += 1
            StrCounter = 'E'+str(index)
            if XL[StrCounter].value != element:
                nextElem = True
                index -= 1
    return index


def BinarySearch(xl, element, lowerbound, upperbound):
    if lowerbound > upperbound:
        return -1
    MidIndex = (lowerbound + upperbound) // 2
    StrCounter = 'E'+str(MidIndex)
    if element > xl[StrCounter].value:
        lowerbound = MidIndex + 1
        return BinarySearch(xl, element, lowerbound, upperbound)
    elif element < xl[StrCounter].value:
        upperbound = MidIndex - 1
        return BinarySearch(xl, element, lowerbound, upperbound)
    return MidIndex


def getStudentsRecord(xl, lowerindex, upperindex):
    Record = []
    SN = 0
    for i in range(lowerindex, upperindex+1):
        SN += 1
        StrCounter = str(i)
        FName = xl['A' + StrCounter].value
        LName = xl['B' + StrCounter].value
        StdID = xl['C' + StrCounter].value
        Email = xl['D' + StrCounter].value
        recordTuple = (SN, FName, LName, StdID, Email)
        Record.append(recordTuple)
    return Record


def BinaryLinearSearch(xl, element):
    Lowerbound = 2
    Upperbound = xl.max_row
    Mid = BinarySearch(xl, element, Lowerbound, Upperbound)
    if Mid != -1:
        LowerIndex = LinearSearch(xl, Mid, element, True)
        UpperIndex = LinearSearch(xl, Mid, element)
        return getStudentsRecord(xl, LowerIndex, UpperIndex)
    return -1


def main(Area):
    Database = openpyxl.load_workbook(
        filename=AbsolutePath('SortedDatabase.xlsx'))
    Datasheet = Database.active
    Records = BinaryLinearSearch(Datasheet, Area)
    if Records == -1:
        return 'No student in your Area'
    return Records
