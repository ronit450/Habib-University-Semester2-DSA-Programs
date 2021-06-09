import openpyxl
def creating_excel_file(final_lst):
    file = openpyxl.load_workbook('Data.xlsx')
    sheet = file.active
    sheet.cell(row=1,column=1).value = 'First name'
    sheet.cell(row=1,column=2).value = 'Last name'
    sheet.cell(row=1,column=3).value = 'Habib Id'
    sheet.cell(row=1,column=4).value = 'Email Id'
    count = 2
    for i in range(len(final_lst)):
        for j in range(len(final_lst[i])):
            sheet.cell(row=count, column=j+1).value = final_lst[i][j]
        count +=1
    file.save('Data.xlsx')
def readFile(name_of_file):
    line_lst = []
    file = open(name_of_file, 'r')
    file_lines = file.readlines()
    for i in file_lines:
        temp = i.strip()
        line_lst.append(temp)
    file.close()
    return line_lst

def helper_of_id(name):
    name = str.lower(name)
    name = name[0:1]
    return name

def id_generator(lst_of_names):
    lst_of_emails = []
    habib_id= []
    code = 1
    for i in lst_of_names:
        name = str.split(i)
        first_letter = helper_of_id(name[0])
        second_letter = helper_of_id(name[1])
        if code <= 9:
            id = first_letter + second_letter +'0600' + str(code) + '@st.hu.edu.pk'
            habib_id.append('0600' + str(code))
        elif code >= 10 and code <= 99:
            id = first_letter + second_letter + '060' + str(code) + '@st.hu.edu.pk'
            habib_id.append('060' + str(code))
        elif code >= 100 and code <= 999:
            id = first_letter + second_letter + '06' + str(code) + '@st.hu.edu.pk'
            habib_id.append('06' + str(code))
        lst_of_emails.append(id)
        code += 1
    return lst_of_emails, habib_id
def main_function():
    list_of_names= readFile('NameChr.txt')
    temp = id_generator(list_of_names)
    list_of_emails = temp[0]
    list_of_habib_id = temp[1]
    temp_lst = []
    for i in range(len(list_of_habib_id)):
        name = str.split(list_of_names[i])
        f_name, lname = name[0], name[1]
        temp_tuple = (f_name, lname, list_of_habib_id[i], list_of_emails[i])
        temp_lst.append(temp_tuple)
    creating_excel_file(temp_lst)
    print(temp_lst)
main_function()


