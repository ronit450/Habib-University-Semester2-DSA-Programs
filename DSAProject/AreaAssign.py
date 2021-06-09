import openpyxl
import random


def readArea():
    Area = openpyxl.load_workbook(filename='DSASurvey.xlsx')
    AreaSheet = Area.active
    AreaList = []
    for i in range(2, AreaSheet.max_row+1):
        temp = 'B' + str(i)
        AreaList.append(AreaSheet[temp].value)
    return AreaList


def getRandomArea(AL):
    Upperbound = len(AL)-1
    Lowerbound = 0
    RandomIndex = random.randint(Lowerbound, Upperbound)
    return AL[RandomIndex]


def main():
    Database = openpyxl.load_workbook(filename='Data.xlsx')
    MainSheet = Database.active
    MainSheet['E1'] = 'Area'
    for i in range(2, MainSheet.max_row+1):
        Col = 'E' + str(i)
        MainSheet[Col] = getRandomArea(readArea())
    Database.save(filename='Database.xlsx')


# main() #Un-comment to run it again and re-create Database.xlsx file
